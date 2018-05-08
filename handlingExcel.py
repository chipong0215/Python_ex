import openpyxl, pprint, logging

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')#, filename='crawlingLog.txt')

logging.info('Opening workbook...')

wb = openpyxl.load_workbook(r'C:\Users\pongl\Documents\Python Scripts\Python_ex\censuspopdata.xlsx')
wb.create_sheet(title='Output')
sheet = wb.get_sheet_by_name('Population by Census Tract')
countyData = {}

logging.info('Reading rows...')

for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts':0, 'pop':0})

    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)

logging.info('Writing result...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()



logging.info('Switching to sheet output...')
output = wb.get_sheet_by_name('Output')
index = 2

# output.cell(row=index, column=1).value = '="State"'
# output.cell(row=index, column=2).value = '="County"'
# output.cell(row=index, column=3).value = '="POP2010"'
# output.cell(row=index, column=4).value = '="CensusTract"'
output.cell(row=1, column=1).value = 'State'
output.cell(row=1, column=2).value = 'County'
output.cell(row=1, column=3).value = 'POP2010'
output.cell(row=1, column=4).value = 'CensusTract'


logging.info('Start to write data...')
for state, county in countyData.items():
    for key, values in county.items():
        # print(k + ':' + x + ':' + str(y['pop']) + ':' + str(y['tracts']))
        output.cell(row=index, column=1).value = state
        output.cell(row=index, column=2).value = key
        output.cell(row=index, column=3).value = values['pop']
        output.cell(row=index, column=4).value = values['tracts']
        index += 1
        

wb.save('text.xlsx')
logging.info('Done.')

