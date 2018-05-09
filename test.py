import openpyxl, pprint, logging, copy

# dict1 = census2010.allData

# # print(dict1)
# # print(type(dict1))

# for k,v in dict1.items():
#     for x,y in v.items():
#         print(k + ':' + x + ':' + str(y['pop']) + ':' + str(y['tracts']))


wb = openpyxl.load_workbook(r'C:\Users\pongl\Documents\Python Scripts\Python_ex\censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('hihi')


path = r'C:\Users\pongl\Documents\Python Scripts\Python_ex\test2.xlsx'
openfile = openpyxl.Workbook()
openfile.save(path)

wb2 = openpyxl.load_workbook(r'C:\Users\pongl\Documents\Python Scripts\Python_ex\test2.xlsx')
sheet2 = wb2.active

# copyrange = sheet['A1':'F4']

for row in list(sheet.rows)[0:4]:
    for col in row:
        tmp = sheet2[col.column + str(col.row)]
        # str1 = str(col.row) + col.column]
        # tmp = sheet2.cell(row=col.row, column=col.column)

        tmp.font = copy.copy(col.font)
        tmp.border = copy.copy(col.border)
        tmp.fill = copy.copy(col.fill)
        tmp.number_format = copy.copy(col.number_format)
        tmp.protection = copy.copy(col.protection)
        tmp.alignment = copy.copy(col.alignment)
        tmp.value = col.value

        if col.column == 'A':
            sheet2.merge_cells(start_row=1,start_column=1,end_row=1,end_column=6)
            continue

        # print(col.row, col.column)

# tmp = sheet2.cell(row=1, column=1)
# if sheet.cell(row=1, column=1).has_style:
#     tmp.font = copy.copy(sheet.cell(row=1, column=1).font)
#     tmp.border = copy.copy(sheet.cell(row=1, column=1).border)
#     tmp.fill = copy.copy(sheet.cell(row=1, column=1).fill)
#     tmp.number_format = copy.copy(sheet.cell(row=1, column=1).number_format)
#     tmp.protection = copy.copy(sheet.cell(row=1, column=1).protection)
#     tmp.alignment = copy.copy(sheet.cell(row=1, column=1).alignment)
#     tmp.value = sheet.cell(row=1, column=1).value


# sheet2.cell(row=1, column=1).value = tmp

wb2.save('test2.xlsx')

# for row in default_sheet.rows:
#     for cell in row:
#         new_cell = new_sheet.cell(row=cell.row_idx,
#                    col=cell.col_idx, value= cell.value)
#         if cell.has_style:
#             new_cell.font = cell.font
#             new_cell.border = cell.border
#             new_cell.fill = cell.fill
#             new_cell.number_format = cell.number_format
#             new_cell.protection = cell.protection
#             new_cell.alignment = cell.alignment


# That would work though it's faster to copy the underlying StyleArray: 

# c1 = ws['A1'] 
# c2 = ws['B2'] 

# c2._style = copy(c1._style) 

# However, what I am working on are NamedStyles. These represent styles that   
# are supposed to be shared just like the "Style Templates" in Excel. The   
# API will be something like this: 

# normal = NamedStyle(name="Normal", font=Font(…), etc.) 
# wb.named_styles.add(normal) 

# ws['A1'].style = "Normal" 

# This would hopefully be easy to work with and also very fast. 

# However, there are a few implementation hurdles to jump: 

#         * named styles and cell formats are, according to the specification,   
# supposed to be commutative so that you could define a named style, apply   
# it to range of cells, change it, add local cell formats and everything   
# would look great. Except Excel explicitly doesn't follow the standard   
# here. Also it's not clear how to resolve possible conflicts. What do you   
# do when a cell font is bold but the named style has font.bold = False?.   
# This is particularly important if the style is supposed to mutable after   
# assignment. 

# The simplest implementation would do sort of what Excel does and copy the   
# attributes of the named style as well as maintaining a link to it but this   
# makes a lot of assumptions about use in code: a named style would need to   
# be immutable after being assigned once; existing cell formats would have   
# to be replaced en masse but could then be overwritten. Or resolution has   
# to be done when serialising the cell, which could slow things down a lot. 

# Also the code for the API is a bit tricky. 
